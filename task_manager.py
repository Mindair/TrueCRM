# task_manager.py

from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import PatternFill
import json
from openpyxl.styles import Alignment  # Импортируем класс Alignment
import uuid  # Импортируем модуль для генерации UUID



class TaskManager:
    def __init__(self, filename="tasks.xlsx"):
        self.tasks = []  # Инициализируем пустой список задач
        self.filename = filename  # Сохраняем имя файла

        try:
            wb = load_workbook(self.filename)
            ws = wb.active

            # Читаем данные из файла
            for row in ws.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовки
                if row[0]:  # Проверяем, что ID указан
                    task = {
                        "id": row[0],  # ID задачи
                        "done": True if row[1] == "Выполнено" else False,
                        "company": row[2],
                        "description": row[3],
                        "deadline": row[4] if row[4] != "Нет дедлайна" else None,
                        "created_at": row[5]
                    }
                    self.tasks.append(task)

        except FileNotFoundError:
            print("Файл задач не найден. Создан новый список задач.")
        except InvalidFileException:
            print("Ошибка чтения файла. Создан новый список задач.")


    def addTask(self):
        description = input("Введите описание задачи: ")
        company = input("Введите название компании: ")
        while True:
            deadline_input = input("Введите дедлайн (YYYY-MM-DD) или нажмите Enter, если нет дедлайна: ").strip()
            if not deadline_input:
                deadline = None
                print("Дедлайн не установлен.")
                break
            else:
                try:
                    if len(deadline_input) != 10 or deadline_input[4] != '-' or deadline_input[7] != '-':
                        raise ValueError("Неверный формат даты. Используйте формат YYYY-MM-DD.")
                    deadline_date = datetime.strptime(deadline_input, '%Y-%m-%d')
                    deadline = deadline_date.strftime('%Y-%m-%d')
                    current_date = datetime.now().date()
                    if deadline_date.date() < current_date:
                        print("Дедлайн должен быть в будущем.")
                        continue
                    print(f"Дедлайн успешно установлен: {deadline}")
                    break
                except ValueError as e:
                    print(f"Ошибка: {e}. Пожалуйста, используйте формат YYYY-MM-DD.")

        # Генерируем уникальный id для задачи
        task_id = str(uuid.uuid4())[:8]  # Берем первые 8 символов UUID для компактности

        # Создаем задачу
        created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # Время создания
        task = {
            "id": task_id,  # Уникальный идентификатор
            "description": description,
            "done": False,
            "company": company,
            "deadline": deadline,
            "created_at": created_at
        }
        self.tasks.append(task)
        print("Задача добавлена!")

    def editTaskDescription(self):
        self.lookTasks()  # Показываем текущие задачи
        if not self.tasks:
            return

        try:
            number = int(input("Введите номер задачи для редактирования описания: "))
            if 1 <= number <= len(self.tasks):
                current_task = self.tasks[number - 1]
                print(f"Текущее описание: {current_task['description']}")

                new_description = input("Введите дополнение к описанию (или нажмите Enter, чтобы пропустить): ").strip()
                if new_description:  # Если пользователь ввел новое описание
                    # Добавляем новое описание через \n
                    current_task["description"] += f"\n{new_description}"
                    print("Описание успешно обновлено.")
                else:
                    print("Описание не изменено.")
            else:
                print("Неверный номер задачи.")
        except ValueError:
            print("Введите корректный номер.")

    def lookTasks(self):
        if not self.tasks:
            print("Список задач пуст.")
            return

        print("\nВсе задачи:")
        for i, task in enumerate(self.tasks, start=1):
            status = "[v]" if task["done"] else "[x]"
            deadline = f"(Дедлайн: {task['deadline']})" if task["deadline"] else ""
            print(
                f"{i}. ID: {task['id']} | {status} Компания: {task['company']} | Задача: {task['description']} {deadline}")

    def removeTask(self):
        self.lookTasks()
        if not self.tasks:
            return
        try:
            number = int(input("Введите номер задачи для удаления: "))
            if 1 <= number <= len(self.tasks):
                removed_task = self.tasks.pop(number - 1)
                print(f"Задача \"{removed_task['description']}\" удалена.")
            else:
                print("Неверный номер задачи.")
        except ValueError:
            print("Введите корректный номер.")

    def markTask(self):
        self.lookTasks()
        if not self.tasks:
            return
        try:
            number = int(input("Введите номер задачи для отметки как выполненной: "))
            if 1 <= number <= len(self.tasks):
                self.tasks[number - 1]["done"] = True
                print("Задача отмечена как выполненная.")
            else:
                print("Неверный номер задачи.")
        except ValueError:
            print("Введите корректный номер.")

    def clearTasks(self):
        confirmation = input("Вы уверены, что хотите удалить ВСЕ задачи? (да/нет): ").strip().lower()
        if confirmation == "да":
            self.tasks = []  # Очищаем список задач
            print("Все задачи успешно удалены.")

            # Сохраняем пустой список в Excel
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Задачи"

                # Обновленные заголовки
                headers = ["Номер", "Статус", "Компания", "Описание", "Дедлайн", "Время создания"]
                ws.append(headers)

                # Сохраняем пустой файл
                wb.save(self.filename)
                print(f"Файл {self.filename} очищен.")
            except Exception as e:
                print(f"Ошибка при очистке файла: {e}")
        else:
            print("Операция отменена.")

    def save_to_json(self, filename="tasks.json"):
        with open(filename, "w", encoding="utf-8") as file:
            json.dump(self.tasks, file, indent=4, ensure_ascii=False)
        print(f"Задачи успешно сохранены в {filename}.")

    def save_to_txt(self, filename="tasks.txt"):
        with open(filename, "w", encoding="utf-8") as file:
            if not self.tasks:
                file.write("Список задач пуст.")
            else:
                for i, task in enumerate(self.tasks, start=1):
                    status = "[v]" if task["done"] else "[x]"
                    deadline = f"(Дедлайн: {task['deadline']})" if task["deadline"] else ""
                    created_at = f"(Создана: {task['created_at']})"  # Выводим время создания
                    file.write(
                        f"{i}. {status} Компания: {task['company']} | Задача: {task['description']} {deadline} {created_at}\n")
        print(f"Задачи успешно сохранены в {filename}.")

    def save_to_excel(self, filename="tasks.xlsx"):
        """Сохраняет задачи в Excel-файл."""
        headers = ["ID", "Статус", "Компания", "Описание", "Дедлайн", "Время создания"]  # Определяем заголовки

        try:
            # Попытка загрузить существующий файл
            wb = load_workbook(filename)
            ws = wb.active
        except FileNotFoundError:
            # Если файл не найден, создаем новый
            wb = Workbook()
            ws = wb.active
            ws.title = "Задачи"
            ws.append(headers)  # Добавляем заголовки

        # Цветовая схема
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                                 fill_type="solid")  # Зеленый для выполненых задач
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                               fill_type="solid")  # Красный для невыполненых задач

        # Словарь для хранения существующих задач по их ID
        existing_tasks = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=len(headers), values_only=True), start=2):
            if row[0]:  # Проверяем, что ID указан
                existing_tasks[row[0]] = {
                    "row_index": row_idx,  # Номер строки в Excel (начиная с 2, так как первая строка — заголовки)
                    "status": row[1],
                    "company": row[2],
                    "description": row[3],
                    "deadline": row[4],
                    "created_at": row[5]
                }

        # Обновляем или добавляем задачи
        for task in self.tasks:
            task_id = task["id"]
            if task_id in existing_tasks:  # Если задача уже существует
                row_index = existing_tasks[task_id]["row_index"]  # Номер строки в Excel

                # Обновляем данные в существующей строке
                ws.cell(row=row_index, column=2, value="Выполнено" if task["done"] else "Не выполнено")
                ws.cell(row=row_index, column=3, value=task["company"])
                ws.cell(row=row_index, column=4, value=task["description"])
                ws.cell(row=row_index, column=5, value=task["deadline"] if task["deadline"] else "Нет дедлайна")

                # Применяем цветовую заливку только к ячейке "Статус"
                status_cell = ws.cell(row=row_index, column=2)  # Ячейка "Статус" (второй столбец)
                if task["done"]:
                    status_cell.fill = green_fill  # Зеленый для выполненых задач
                else:
                    status_cell.fill = red_fill  # Красный для невыполненых задач

                # Форматирование ячейки "Описание"
                description_cell = ws.cell(row=row_index, column=4)  # Ячейка "Описание" (четвертый столбец)
                description_cell.alignment = Alignment(wrap_text=True)  # Включаем перенос текста
            else:  # Если задачи нет в файле, добавляем новую строку
                row = [
                    task["id"],  # ID задачи
                    "Выполнено" if task["done"] else "Не выполнено",
                    task["company"],
                    task["description"],
                    task["deadline"] if task["deadline"] else "Нет дедлайна",
                    task["created_at"]
                ]
                ws.append(row)

                # Применяем цветовую заливку только к ячейке "Статус"
                row_index = ws.max_row  # Номер текущей строки
                status_cell = ws.cell(row=row_index, column=2)  # Ячейка "Статус" (второй столбец)
                if task["done"]:
                    status_cell.fill = green_fill  # Зеленый для выполненых задач
                else:
                    status_cell.fill = red_fill  # Красный для невыполненых задач

                # Форматирование ячейки "Описание"
                description_cell = ws.cell(row=row_index, column=4)  # Ячейка "Описание" (четвертый столбец)
                description_cell.alignment = Alignment(wrap_text=True)  # Включаем перенос текста

        # Сохраняем файл
        wb.save(filename)
        print(f"Задачи успешно сохранены в {filename}.")