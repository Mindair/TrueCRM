# file_savers.py

import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def save_to_json(tasks, filename="tasks.json"):
    with open(filename, "w", encoding="utf-8") as file:
        json.dump(tasks, file, indent=4, ensure_ascii=False)
    print(f"Задачи успешно сохранены в {filename}.")


def save_to_txt(tasks, filename="tasks.txt"):
    with open(filename, "w", encoding="utf-8") as file:
        if not tasks:
            file.write("Список задач пуст.")
        else:
            for i, task in enumerate(tasks, start=1):
                status = "[v]" if task["done"] else "[x]"
                deadline = f"(Дедлайн: {task['deadline']})" if task["deadline"] else ""
                file.write(f"{i}. {status} Компания: {task['company']} | Задача: {task['description']} {deadline}\n")
    print(f"Задачи успешно сохранены в {filename}.")


def save_to_excel(tasks, filename="tasks.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Задачи"

    headers = ["Номер", "Статус", "Компания", "Описание", "Дедлайн"]
    ws.append(headers)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for i, task in enumerate(tasks, start=1):
        status = "Выполнено" if task["done"] else "Не выполнено"
        deadline = task["deadline"] if task["deadline"] else "Нет дедлайна"
        row = [i, status, task["company"], task["description"], deadline]
        ws.append(row)

        row_index = ws.max_row
        if task["done"]:
            for cell in ws[row_index]:
                cell.fill = green_fill
        else:
            for cell in ws[row_index]:
                cell.fill = red_fill

    wb.save(filename)
    print(f"Задачи успешно сохранены в {filename}.")